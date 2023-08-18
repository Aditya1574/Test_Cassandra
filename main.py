import requests
import openpyxl
import string
from cassandra.cluster import Cluster

#
cluster = Cluster()
session = cluster.connect('test_ks')

def read_specific_columns(filename, sheetname, columns, start_row, num_rows):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]

    rows_data = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if row_num > start_row + num_rows - 1:
            break

        selected_columns = [row[column_index] for column_index in columns]
        rows_data.append(selected_columns)

    return rows_data

def column_index(column_letter):
    return string.ascii_uppercase.index(column_letter)


def read_excel_column(filename, sheetname, column_letter):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]

    column_data = []
    column_idx = column_index(column_letter)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        column_data.append(row[column_idx])

    return column_data

def read_excel_rows(filename, sheetname, start_row, num_rows):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]

    rows_data = []
    for row in sheet.iter_rows(min_row=start_row, max_row=start_row + num_rows - 1, values_only=True):
        rows_data.append(row)

    return rows_data

def get_weather_data(api_key, location):
    base_url = "http://api.weatherstack.com/current"

    params = {
        'access_key': api_key,
        'query': location
    }

    try:
        response = requests.get(base_url, params=params)
        response.raise_for_status()  # Raise an exception for HTTP errors
        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        print("Error:", e)
        return None


def main():
    api_key = '864619f072761438aa67cb735790ba50'
    num_rows = 5
    start_row = 2
    filename = './Data_Sources/MOCK_DATA.xlsx'
    sheetname = 'data'
    column_letter = 'F'

    column_array = read_specific_columns(filename, sheetname, [5], start_row, num_rows)  # getting the data of the employee
    employee_weather_data = []
    # extracting weather data
    for cityName in column_array:
        location = cityName
        # 'New York'  # Example location

        weather_data = get_weather_data(api_key, location)

        if weather_data:
            temperature = weather_data['current']['temperature']
            weather_description = weather_data['current']['weather_descriptions'][0]

            employee_weather_data.append((temperature,weather_description))


    employee_rows = []
    print("Select the rows from the emloyees table \n")
    all_rows = read_excel_rows(filename,sheetname,1,1)[0]
    employee_dict = dict()
    for i in range(len(all_rows)):
        print(f"{i+1} {all_rows[i]} \n")
        employee_dict[i] = all_rows[i]
    q = 1
    while q == 1:
        element_selected  = int(input("Enter the choice \n"))
        employee_rows.append(element_selected - 1)
        q = int(input("Select more ? q = (1/0) \n"))
    weather_dict = dict()
    print("Select the rows from the weather data \n")
    print("1. Temperature \n 2. Conditions \n")
    weather_dict[0] = "Temperature"
    weather_dict[1] = "Conditions"
    q = 1
    weather_rows = []
    while q == 1:
        element_selected  = int(input("Enter the choice \n"))
        weather_rows.append(element_selected - 1)
        q = int(input("Select more ? q = (1/0) \n"))
    # Debugging statements
    # print(employee_rows)
    # print("\n")
    # print(employee_dict)
    # print("\n")
    # print(weather_rows)
    # print("\n")
    # print(weather_dict)
    # print("\n")
    primary_key = input("What is the primary Key ? \n")
    table_name  = input("Enter table name ? \n")
    column_name = ""
    column_name_addValues = ""
    print(employee_dict)
    datatype_mapper = {"id" : "int", "first_name" : "str", "last_name" : "str", "email" : "str", "gender" : "str", "City" : "str"}
    for i in range(len(employee_rows)):
        col_name = employee_dict[employee_rows[i]]
        col_name_addValues = employee_dict[employee_rows[i]]
        col_name += " int" if datatype_mapper[col_name] == "int" else " text"

        if employee_dict[employee_rows[i]] == primary_key:
            col_name = col_name + " PRIMARY KEY"
        if i == len(employee_rows) - 1:
            column_name = column_name + col_name
            column_name_addValues = column_name_addValues + col_name_addValues
        else:
            column_name = column_name + col_name + ","
            column_name_addValues = column_name_addValues + col_name_addValues +  ","

    if 0 in weather_rows:
        column_name = column_name + "," + weather_dict[0] + " int"
    if 1 in weather_rows:
        column_name = column_name + "," + weather_dict[1] + " text"

    create_query = f"CREATE TABLE {table_name}({column_name});"
    print(create_query)
    session.execute(create_query)

    if 0 in weather_rows:
        column_name_addValues = column_name_addValues + "," + weather_dict[0]
    if 1 in weather_rows:
        column_name_addValues = column_name_addValues + "," + weather_dict[1]

    rows_array = read_specific_columns(filename, sheetname, employee_rows, start_row, num_rows)
    insert_comm = f"INSERT INTO {table_name} ({column_name_addValues}) "

    for i in range(len(rows_array)):
        value_comm = "VALUES("
        for j in range(len(rows_array[i])):
                if isinstance(rows_array[i][j], str):
                    value_comm += "" if j == 0 else ","
                    value_comm += "'" + rows_array[i][j] + "'"
                else:
                    value_comm += "" if j == 0 else ","
                    value_comm += str(rows_array[i][j])
        if 0 in weather_rows:
            value_comm = value_comm + ","  + str(employee_weather_data[i][0])
        if 1 in weather_rows:
            value_comm = value_comm + "," + "'" + employee_weather_data[i][1] + "'"

        value_comm = value_comm + ");"

        query = insert_comm + value_comm
        print(query)
        session.execute(query)
        print(f"inserted row {i +1}")
        print("\n")



if __name__ == "__main__":
    main()
